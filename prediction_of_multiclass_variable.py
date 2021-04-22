import time

start = time.time()
from keras.models import load_model
import cv2
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

import random
import gc

from sklearn.model_selection import train_test_split
from keras import layers
from keras import models
from keras import optimizers
from keras.preprocessing.image import ImageDataGenerator
from keras.preprocessing.image import img_to_array, load_img
from keras.models import load_model
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from pydub import AudioSegment
import librosa
import pydub
import contextlib
import wave
from os import listdir
import os
# import audiosegment
import numpy as np
import matplotlib.pyplot as plt
import itertools
import librosa.display
from sys import byteorder
from array import array
from struct import pack

from scipy import signal
from scipy.io import wavfile
import random
from skimage.measure import block_reduce
import time
from keras import models
from keras import layers
from skimage.transform import resize

import datetime
import scipy as sp
from scipy.io.wavfile import read
from scipy.io.wavfile import write

import sys
import math

import os.path
from os import path

main_path_list = ['C:/Users/arunadevi.r/Desktop/fin_testt',
       'C:/Users/arunadevi.r/Desktop/fin_test_work_wav/',
       'C:/Users/arunadevi.r/Desktop/fin_test/',
       'C:/Users/arunadevi.r/Desktop/1sec_fin_test',
       'C:/Users/arunadevi.r/Desktop/1sec_fin_test_image',
       'C:/Users/arunadevi.r/Desktop/mmuus/(90)eiep_ho_best_model.h5']

model = load_model(main_path_list[5])


def process_and_labeling(img):
    X = []
    for image in img:
        X.append(cv2.resize(cv2.imread(image, cv2.IMREAD_COLOR), (nrows, ncloumns), interpolation=cv2.INTER_CUBIC))
    return X


def split_function(start_time, end_time, file):
    frrames = []
    so = AudioSegment.from_wav(file)
    print('so', so)
    with contextlib.closing(wave.open(file, 'r')) as f:
        firames = f.getnframes()
        irate = f.getframerate()
        iduration = firames / float(irate)
    ss = start_time
    sec_ee = 0
    lllpo = []
    while sec_ee <= iduration:
        ee = ss + end_time
        lllpo.append([ss, ee])
        frrames.append(so[ss:ee])
        ss = ee
        drf = ee / 1000
        sec_ee = drf
    return frrames, lllpo


def findDuration(fname):
    with contextlib.closing(wave.open(fname, 'r')) as f:
        frames = f.getnframes()
        rate = f.getframerate()
        sw = f.getsampwidth()
        chan = f.getnchannels()
        duration = frames / float(rate)
        # print("File:", fname, "--->",frames, rate, sw, chan)
        return duration


def convert(seconds):
    seconds = seconds % (24 * 3600)
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
    return "%d:%02d:%02d" % (hour, minutes, seconds)


paths = [main_path_list[0]]
wavpath = [main_path_list[1]]
for l in range(len(paths)):
    fname = [f for f in listdir(paths[l])]
    print('fname', fname)
    for wavb in fname:
        print('wavb', wavb)
        # outname = 'filtered.wav'
        cutOffFrequency = 400.0
        print(paths[l] + '/' + wavb)
        auddio = AudioSegment.from_wav(paths[l] + '/' + wavb)
        auddio.export(main_path_list[1] + 'working_wav.wav', format='wav')


        def running_mean(x, windowSize):
            cumsum = np.cumsum(np.insert(x, 0, 0))
            return (cumsum[windowSize:] - cumsum[:-windowSize]) / windowSize


        def interpret_wav(raw_bytes, n_frames, n_channels, sample_width, interleaved=True):
            if sample_width == 1:
                dtype = np.uint8
            elif sample_width == 2:
                dtype = np.int16
            else:
                raise ValueError("Only supports 8 and 16 bit audio formats.")
            channels = np.fromstring(raw_bytes, dtype=dtype)
            if interleaved:
                channels.shape = (n_frames, n_channels)
                channels = channels.T
            else:
                channels.shape = (n_channels, n_frames)
            return channels


        with contextlib.closing(wave.open(main_path_list[1] + 'working_wav.wav', 'rb')) as spf:
            sampleRate = spf.getframerate()
            ampWidth = spf.getsampwidth()
            nChannels = spf.getnchannels()
            nFrames = spf.getnframes()
            signal = spf.readframes(nFrames * nChannels)
            spf.close()
            channels = interpret_wav(signal, nFrames, nChannels, ampWidth, True)
            freqRatio = (cutOffFrequency / sampleRate)
            N = int(math.sqrt(0.196196 + freqRatio ** 2) / freqRatio)
            filtered = running_mean(channels[0], N).astype(channels.dtype)
            wav_file = wave.open(main_path_list[2] + 'filtered.wav', "w")
            wav_file.setparams((1, ampWidth, sampleRate, nFrames, spf.getcomptype(), spf.getcompname()))
            wav_file.writeframes(filtered.tobytes('C'))
            wav_file.close()

        filtered_file_path = [main_path_list[2]]

        for l in range(len(filtered_file_path)):
            number_of_image = 0
            filtered_file_name = [f for f in listdir(filtered_file_path[l])]
            print('split', filtered_file_path[l] + '/' + filtered_file_name[0])
            # print('over_all_file_in_path',len(name_ff))
            for file_name in filtered_file_name:
                print('frame', filtered_file_path[l] + '/' + file_name)
                frame, times = split_function(0, 1000, filtered_file_path[l] + '/' + file_name)
                # print(len(frame))
                for file_number in range(len(frame) - 1):
                    int_file_number = str(file_number)
                    signle_channel_file = frame[file_number].set_channels(1)
                    signle_channel_file.export(main_path_list[3] + '/' + '1_sec_frame' + int_file_number + '.wav', format='wav')

        split_path = [main_path_list[3]]
        spectrogram_image_path = [main_path_list[4]]

        number_of_image = 0
        for split_audio in range(len(split_path)):
            number_of_image = 0
            filtered_file_name = [f for f in listdir(split_path[split_audio])]

            for filtered_file in range(len(filtered_file_name)):
                integer_filter_filenumber = str(filtered_file)

                audio_time_series, sample_rate = librosa.load(split_path[l] + '/' + filtered_file_name[filtered_file])
                magnitude, phase = librosa.magphase(librosa.stft(audio_time_series))
                fig_size = plt.figure(figsize=(2, 2))
                librosa.display.specshow(librosa.amplitude_to_db(magnitude[:], ref=np.max), sr=sample_rate)
                plt.tight_layout()

                plt.savefig(spectrogram_image_path[l] + '/' + integer_filter_filenumber + 'single_sec')
                plt.close()

        spectrogram_imagepath = main_path_list[4]
        spectrogram_for_prediction = [spectrogram_imagepath + '/' + f for f in listdir(spectrogram_imagepath)]

        nrows = 150
        ncloumns = 150
        channels = 3

        X_test = process_and_labeling(spectrogram_for_prediction)
        pre_b_num = len(X_test)
        del spectrogram_for_prediction

        X_test = np.array(X_test)

        test_gen = ImageDataGenerator(rescale=1. / 255)

        # file_names=[f for f in listdir(pat[3])]

        i = 0
        label = []
        for bb in test_gen.flow(X_test, batch_size=1):
            prob = model.predict(bb)
            pred = np.argmax(prob, axis=-1)
            label.append(pred)
            i += 1
            if i % pre_b_num == 0:
                break

        # segment_list=[]
        # for ggl in range(len(file_names)):
        #   segment=AudioSegment.from_file(pat[3]+'/'+file_names[ggl])
        #  segment_list.append(segment)
        # os.remove(pat[3]+'/'+file_names[ggl])

        # print(len(segment_list))

        integer_filter_filenumber = label
        ind = pd.DataFrame(data=integer_filter_filenumber)
        ind.columns = ['pre']
        # print('label',s)
        ivr = ind[ind['pre'] == 0]
        rep = ind[ind['pre'] == 1]
        client = ind[ind['pre'] == 2]
        music = ind[ind['pre'] == 3]
        noise = ind[ind['pre'] == 4]

        f_r = []
        lis_lis = []

        f_r.append(wavb)

        music_total_sec = len(music)
        if music_total_sec:
            f_r.append(convert(music_total_sec))
        else:
            print('no music')
            f_r.append('no music')

        ivr_total_sec = len(ivr)
        if ivr_total_sec:
            f_r.append(convert(ivr_total_sec))
        else:
            print('IVR System voice')
            f_r.append('IVR System voice')

        rep_total_sec = len(rep)
        if rep_total_sec:
            f_r.append(convert(rep_total_sec))
        else:
            print('AHS Client Partner')
            f_r.append('AHS Client Partner')

        client_total_sec = len(client)
        if client_total_sec:
            f_r.append(convert(client_total_sec))
        else:
            print('no Payer Operator')
            f_r.append('no Payer Operator')

        noise_total_sec = len(noise)
        if noise_total_sec:
            f_r.append(convert(noise_total_sec))
        else:
            print('no slience')
            f_r.append('no slience')

        t_d = findDuration(main_path_list[1] + 'working_wav.wav')

        f_r.append(convert(t_d))

        lis_lis.append(f_r)

        excel_ex = path.exists('C:/Users/sarathkumar.h/Desktop/mp33/test.xlsx')

        if excel_ex == False:
            headers = ['file_name', 'music_duration', 'IVR System voice duration', 'AHS Client Partner duration',
                       'Payer Operator duration', 'slience duration', 'total_duration_of_file']
            workbook_name = 'C:/Users/sarathkumar.h/Desktop/mp33/test.xlsx'
            wb = Workbook()
            page = wb.active
            page.title = 'test_result'
            page.append(headers)
            page.append(lis_lis[0])
            wb.save(filename=workbook_name)
            f_r.clear()
            lis_lis.clear()
            label.clear()



        else:
            workbook_name = 'C:/Users/sarathkumar.h/Desktop/mp33/test.xlsx'
            wb = load_workbook(workbook_name)
            page = wb.active
            page.append(lis_lis[0])
            wb.save(filename=workbook_name)
            f_r.clear()
            lis_lis.clear()
            label.clear()

        os.remove(main_path_list[2] + 'filtered.wav')
        deleting_1sec_fin_test = [os.remove(main_path_list[3] + '/' + f) for f in listdir(main_path_list[3])]
        deleting_1sec_fin_test_image = [os.remove(main_path_list[4] + '/' + f) for f in listdir(main_path_list[4])]
        os.remove(main_path_list[1] + 'working_wav.wav')

end = time.time()
print(end - start)